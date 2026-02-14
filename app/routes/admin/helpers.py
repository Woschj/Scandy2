from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, g, send_file, current_app, abort, session
from flask_login import current_user
from app.utils.decorators import admin_required, mitarbeiter_required, login_required
from app.models.mongodb_models import MongoDBUser
from werkzeug.utils import secure_filename
import os
from pathlib import Path
import json
import colorsys
import logging
from datetime import datetime, timedelta
from app.utils.backup_manager import backup_manager
import openpyxl
from io import BytesIO
import time
from PIL import Image
from app.config.config import Config
from app.models.mongodb_database import mongodb
from app.models.feature_system import feature_system, get_feature_settings, set_feature_setting, is_feature_enabled
from app.services.admin_system_settings_service import AdminSystemSettingsService
from app.models.mongodb_models import MongoDBTool, MongoDBWorker, MongoDBConsumable
from bson import ObjectId
from werkzeug.security import generate_password_hash, check_password_hash
from app.utils.database_helpers import get_categories_from_settings, get_ticket_categories_from_settings, get_departments_from_settings, get_locations_from_settings
from docxtpl import DocxTemplate
from urllib.parse import unquote
from werkzeug.utils import secure_filename
import tempfile
import os
import pandas as pd
import tempfile
from typing import Union
import re

# Import der neuen Services
from app.services.admin_dashboard_service import AdminDashboardService
from app.services.admin_user_service import AdminUserService
from app.services.admin_backup_service import AdminBackupService
from app.services.admin_system_service import AdminSystemService
from app.services.admin_email_service import AdminEmailService
from app.services.admin_notification_service import AdminNotificationService
from app.services.admin_ticket_service import AdminTicketService
from app.services.admin_debug_service import AdminDebugService
from app.services.admin_system_settings_service import AdminSystemSettingsService
from app.services.excel_export_service import ExcelExportService
from app.services.custom_fields_service import CustomFieldsService
from app.utils.permissions import (
    get_role_permissions,
    set_role_permissions,
    DEFAULT_ROLE_PERMISSIONS,
    ALLOWED_ACTIONS,
    get_all_actions,
    normalize_permissions,
)

from app.utils.id_helpers import convert_id_for_query, find_document_by_id, find_user_by_id

# Logger einrichten
logger = logging.getLogger(__name__)


# Stelle sicher, dass die Standard-Einstellungen beim Start der App vorhanden sind
def ensure_default_settings():
    """Stellt sicher, dass die Standard-Label-Einstellungen vorhanden sind"""
    default_settings = [
        {'key': 'label_tickets_name', 'value': 'Tickets'},
        {'key': 'label_tickets_icon', 'value': 'fas fa-ticket-alt'},
        {'key': 'label_tools_name', 'value': 'Werkzeuge'},
        {'key': 'label_tools_icon', 'value': 'fas fa-tools'},
        {'key': 'label_consumables_name', 'value': 'Verbrauchsmaterial'},
        {'key': 'label_consumables_icon', 'value': 'fas fa-box-open'}
    ]

    for setting in default_settings:
        mongodb.update_one('settings',
                         {'key': setting['key']},
                         {'$setOnInsert': setting},
                         upsert=True)

def create_excel(data, columns):
    """Erstellt eine Excel-Datei aus Daten"""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Header
    for col, header in enumerate(columns, 1):
        ws.cell(row=1, column=col, value=header)

    # Daten
    for row, item in enumerate(data, 2):
        for col, key in enumerate(columns.keys(), 1):
            value = item.get(key, '')
            ws.cell(row=row, column=col, value=value)

    # Speichern
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def create_multi_sheet_excel(data_dict):
    """Erstellt eine Excel-Datei mit mehreren Arbeitsblättern"""
    wb = openpyxl.Workbook()

    # Entferne das Standard-Arbeitsblatt
    wb.remove(wb.active)

    for sheet_name, data in data_dict.items():
        ws = wb.create_sheet(title=sheet_name)

        if data and len(data) > 0:
            # Spezielle Behandlung für Tools und Consumables
            if sheet_name == 'Werkzeuge':
                _create_enhanced_tools_sheet(ws, data)
            elif sheet_name == 'Verbrauchsmaterial':
                _create_enhanced_consumables_sheet(ws, data)
            else:
                # Standard-Behandlung für andere Sheets
                headers = list(data[0].keys())

                # Header schreiben
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                # Daten schreiben
                for row, item in enumerate(data, 2):
                    for col, key in enumerate(headers, 1):
                        value = item.get(key, '')
                        ws.cell(row=row, column=col, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def _resolve_user_group_names(group_ids):
    """Löst Nutzergruppen-IDs zu Namen auf"""
    from app.utils.id_helpers import resolve_user_group_names
    return resolve_user_group_names(group_ids)

def _create_enhanced_tools_sheet(ws, tools_data):
    """Erstellt eine erweiterte Tools-Tabelle mit allen Feldern"""
    try:
        # Lade Custom Fields für dynamische Header
        try:
            from app.services.custom_fields_service import CustomFieldsService
            custom_fields = CustomFieldsService.get_custom_fields_for_target('tools')
        except Exception as e:
            logger.warning(f"Fehler beim Laden der Custom Fields für Tools: [Interner Fehler]")
            custom_fields = []

        # Header definieren
        headers = [
            'barcode', 'name', 'category', 'location', 'description', 'status',
            'serial_number', 'invoice_number', 'mac_address', 'mac_address_wlan',
            'user_groups', 'additional_software', 'created_at', 'updated_at'
        ]

        # Custom Fields Header hinzufügen - speichere Mapping
        custom_field_mapping = {}
        for custom_field in custom_fields:
            header_name = custom_field['name']
            headers.append(header_name)
            custom_field_mapping[header_name] = custom_field['field_key']

        # Header schreiben
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Daten schreiben
        for row, tool in enumerate(tools_data, 2):
            for col, key in enumerate(headers, 1):
                if key in custom_field_mapping:
                    # Custom Field Wert
                    field_key = custom_field_mapping[key]
                    tool_custom_fields = tool.get('custom_fields', {})
                    value = tool_custom_fields.get(field_key, '')

                    # Formatierung für Custom Fields
                    if isinstance(value, bool):
                        value = 'Ja' if value else 'Nein'
                    elif value is None:
                        value = ''
                elif key == 'user_groups':
                    # Nutzergruppen formatieren - Namen statt IDs
                    groups = tool.get('user_groups', [])
                    value = _resolve_user_group_names(groups)
                elif key == 'additional_software':
                    # Software formatieren
                    software = tool.get('additional_software', [])
                    value = ', '.join(software) if software else ''
                else:
                    # Standard-Wert
                    value = tool.get(key, '')

                ws.cell(row=row, column=col, value=value)
    except Exception as e:
        logger.error(f"Fehler beim Erstellen der erweiterten Tools-Tabelle: [Interner Fehler]")

def _create_enhanced_consumables_sheet(ws, consumables_data):
    """Erstellt eine erweiterte Consumables-Tabelle mit allen Feldern"""
    try:
        # Lade Custom Fields für dynamische Header
        try:
            from app.services.custom_fields_service import CustomFieldsService
            custom_fields = CustomFieldsService.get_custom_fields_for_target('consumables')
        except:
            custom_fields = []

        # Header definieren
        headers = [
            'barcode', 'name', 'category', 'location', 'description', 'quantity',
            'min_quantity', 'created_at', 'updated_at'
        ]

        # Custom Fields Header hinzufügen - speichere Mapping
        custom_field_mapping = {}
        for custom_field in custom_fields:
            header_name = custom_field['name']
            headers.append(header_name)
            custom_field_mapping[header_name] = custom_field['field_key']

        # Header schreiben
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Daten schreiben
        for row, consumable in enumerate(consumables_data, 2):
            for col, key in enumerate(headers, 1):
                if key in custom_field_mapping:
                    # Custom Field Wert
                    field_key = custom_field_mapping[key]
                    consumable_custom_fields = consumable.get('custom_fields', {})
                    value = consumable_custom_fields.get(field_key, '')

                    # Formatierung für Custom Fields
                    if isinstance(value, bool):
                        value = 'Ja' if value else 'Nein'
                    elif value is None:
                        value = ''
                else:
                    # Standard-Wert
                    value = consumable.get(key, '')

                ws.cell(row=row, column=col, value=value)
    except Exception as e:
        logger.error(f"Fehler beim Erstellen der erweiterten Consumables-Tabelle: [Interner Fehler]")

from .blueprint import bp
